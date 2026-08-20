"""Extracts a single sheet (or a row range within one) from a source
workbook into its own standalone .xlsx file, preserving the original
formatting (cell styles, merged ranges, column widths, row heights) --
unlike table_export.py, which writes a clean re-flattened version from
the extractor's parsed columns.

This exists because a dataset's "download" should give back the actual
government table as published, not a reformatted reconstruction.

When multiple tables share one physical sheet (e.g. an urban/rural pair
back to back), pass row_start/row_end (1-indexed, inclusive -- matches
the "sheet_row_start"/"sheet_row_end" extractor.py puts on each table)
so each dataset's download contains only its own table, not its
sheet-mates'.
"""

import copy
from io import BytesIO
from typing import Optional

import openpyxl


def extract_sheet_with_formatting(
    wb, sheet_name: str, row_start: Optional[int] = None, row_end: Optional[int] = None
) -> bytes:
    source_ws = wb[sheet_name]
    r_start = row_start or 1
    r_end = row_end or source_ws.max_row
    row_offset = r_start - 1  # so the export starts at row 1, not mid-sheet

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = str(sheet_name)[:31] or "Sheet1"

    for row in source_ws.iter_rows(min_row=r_start, max_row=r_end):
        for cell in row:
            new_cell = new_ws.cell(
                row=cell.row - row_offset, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    for merged_range in source_ws.merged_cells.ranges:
        # Only carry over merges fully contained in this table's rows --
        # a merge spanning into a sibling table's rows would corrupt the
        # export.
        if merged_range.min_row >= r_start and merged_range.max_row <= r_end:
            shifted = openpyxl.worksheet.cell_range.CellRange(
                min_col=merged_range.min_col, max_col=merged_range.max_col,
                min_row=merged_range.min_row - row_offset,
                max_row=merged_range.max_row - row_offset,
            )
            new_ws.merge_cells(str(shifted))

    for col_letter, dim in source_ws.column_dimensions.items():
        if dim.width:
            new_ws.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in source_ws.row_dimensions.items():
        if dim.height and r_start <= row_idx <= r_end:
            new_ws.row_dimensions[row_idx - row_offset].height = dim.height

    buf = BytesIO()
    new_wb.save(buf)
    return buf.getvalue()


def extract_sheet_with_formatting_from_bytes(
    file_bytes: bytes, sheet_name: str, row_start: Optional[int] = None, row_end: Optional[int] = None
) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    return extract_sheet_with_formatting(wb, sheet_name, row_start, row_end)
