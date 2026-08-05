import concurrent.futures
import io
import json
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import anthropic
from anthropic import RateLimitError

ANTHROPIC_MODEL = "claude-haiku-4-5-20251001"

DDI_PREFIX = "DDI_DEL_DES_VS"
DDI_YEAR   = "2024"
DDI_VER    = "V1"


def _row_text(row) -> str:
    """Join non-empty cells of a raw header row into a single string."""
    if isinstance(row, (list, tuple)):
        return " ".join(str(c).strip() for c in row if c is not None and str(c).strip())
    return str(row).strip()


def _build_ddi_id(tbl: dict) -> str:
    """
    Build a DDI-format table ID.

    _strip_title_desc puts the raw title rows into tbl["title"] / tbl["description"];
    raw_header_rows holds the COLUMN header rows (not the table label rows).

    Mapping:
      tbl["title"]         e.g. "Table : D-3"                                  → code "D3"
      tbl["description"]   e.g. "Live Birth by Age … (Rural)"                  → top  "RURAL"
      raw_header_rows[0]   e.g. ["Religion-All", None, …]  (optional sub-row)  → nxt  "ALL"

    Result: DDI_DEL_DES_VS_B14_RURAL_ALL_2024_V1
    """
    title       = tbl.get("title", "")
    description = tbl.get("description", "")
    raw_headers = tbl.get("raw_header_rows", [])

    # 1. Table code from title row: "Table : D-3" → "D3", "Table : B-14" → "B14"
    code = ""
    m = re.search(r"\b([A-Za-z]-\d+(?:\.\d+)?)\b", title)
    if m:
        code = re.sub(r"[^A-Z0-9]", "", m.group(1).upper())

    # 2. Top-level from the LAST parenthetical in description: "(Urban)" → "URBAN"
    #    The geographic/scope qualifier is always the last parenthetical.
    #    Fall back to title if description has none.
    top = ""
    for src in [description, title]:
        matches = re.findall(r"\(([^)]+)\)", src)
        if matches:
            candidate = re.sub(r"[^A-Z0-9]", "", matches[-1].strip().upper())
            if candidate and not candidate.isdigit():
                top = candidate
                break

    # 3. Next-level from first raw_header_rows entry: "Religion-All" → "ALL"
    #    Looks for "Word-Value" pattern at end of the row text.
    nxt = ""
    for row in raw_headers:
        text = _row_text(row)
        m = re.search(r"[A-Za-z][A-Za-z\s]*-\s*([A-Za-z][A-Za-z\s]*)$", text.strip())
        if m:
            candidate = re.sub(r"[^A-Z0-9]", "", m.group(1).strip().upper())
            if candidate:
                nxt = candidate
                break

    parts = [DDI_PREFIX]
    if code:
        parts.append(code)
    if top:
        parts.append(top)
    if nxt:
        parts.append(nxt)
    parts.append(DDI_YEAR)
    parts.append(DDI_VER)

    return "_".join(parts)


def _parse_retry_delay(error: RateLimitError) -> Optional[float]:
    """
    Extract the server-suggested retry delay from a 429 error body.
    Gemini embeds a retryDelay field (e.g. "24s") in the error details.
    Returns seconds as a float, or None if not found.
    """
    try:
        body = getattr(error, "body", None) or {}
        for detail in body.get("error", {}).get("details", []):
            raw = detail.get("retryDelay", "")
            if raw:
                m = re.match(r"([\d.]+)", str(raw))
                if m:
                    return float(m.group(1)) + 3  # add 3 s buffer
    except Exception:
        pass
    return None


def _is_daily_quota_error(error: RateLimitError) -> bool:
    """
    Return True when the quota that's exhausted resets daily (not per-minute).
    Retrying in the same session won't help — fail fast so the caller falls
    back to heuristic immediately instead of burning 60+ seconds waiting.
    """
    try:
        body = getattr(error, "body", None) or {}
        for detail in body.get("error", {}).get("details", []):
            for v in detail.get("violations", []):
                if "PerDay" in v.get("quotaId", ""):
                    return True
        # Also check the plain message string as a fallback
        msg = body.get("error", {}).get("message", "")
        if "PerDay" in msg or "per_day" in msg.lower():
            return True
    except Exception:
        pass
    return False


def _call_with_retry(client: anthropic.Anthropic, max_retries: int = 3, **kwargs):
    """
    Call client.messages.create with smart backoff on RateLimitError.
    - Checks Retry-After header for the suggested wait time.
    - Falls back to 30 s / 60 s exponential backoff when no header is present.
    """
    for attempt in range(max_retries):
        try:
            return client.messages.create(**kwargs)
        except RateLimitError as e:
            if attempt == max_retries - 1:
                raise
            wait = _parse_retry_delay(e) or (30 * (2 ** attempt))
            print(f"Rate limit — waiting {wait:.0f}s before retry {attempt + 2}/{max_retries} …")
            time.sleep(wait)


def _extract_json_with_key(text: str, required_key: str) -> Optional[Dict]:
    """Find the first valid JSON object in `text` that contains `required_key`."""
    # Strip markdown fences first
    text = re.sub(r"```[a-z]*\n?", "", text).strip().rstrip("`")
    decoder = json.JSONDecoder()
    idx = 0
    while idx < len(text):
        # Find the next opening brace
        start = text.find("{", idx)
        if start == -1:
            break
        try:
            obj, end_pos = decoder.raw_decode(text, start)
            if isinstance(obj, dict) and required_key in obj:
                return obj
            idx = start + 1
        except json.JSONDecodeError:
            idx = start + 1
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# TableExtractor
# ═══════════════════════════════════════════════════════════════════════════════

class TableExtractor:
    def __init__(self, api_key: Optional[str] = None):
        self.client = anthropic.Anthropic(api_key=api_key)

    # ── Public API ────────────────────────────────────────────────────────────

    def extract_from_file(self, file_content: bytes, filename: str) -> List[Dict]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Cannot open Excel file: {e}")

        # Process sheets concurrently -- each sheet is 1+ independent LLM calls,
        # and a large workbook (e.g. 36 sheets) run sequentially is the main
        # source of slowness. Capped at 5 workers to stay well under Anthropic
        # rate limits; _call_with_retry still backs off if we hit them anyway.
        sheets = [(name, wb[name]) for name in wb.sheetnames]
        results_by_sheet: List[List[Dict]] = [[] for _ in sheets]
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as pool:
            future_to_idx = {
                pool.submit(self._process_sheet, ws, name, filename): i
                for i, (name, ws) in enumerate(sheets)
            }
            for future in concurrent.futures.as_completed(future_to_idx):
                results_by_sheet[future_to_idx[future]] = future.result()

        all_tables: List[Dict] = []
        for tables in results_by_sheet:
            all_tables.extend(tables)

        # Generate DDI-format table ID from raw header rows.
        # Format: DDI_DEL_DES_VS_{TABLECODE}_{TOPLEVEL}[_{NEXTLEVEL}]_2024_V1
        for tbl in all_tables:
            tbl["id"] = _build_ddi_id(tbl)

        # Deduplicate IDs within this batch (append _2, _3 for collisions)
        seen_ids: Dict[str, int] = {}
        for tbl in all_tables:
            base = tbl["id"]
            if base in seen_ids:
                seen_ids[base] += 1
                tbl["id"] = f"{base}_{seen_ids[base]}"
            else:
                seen_ids[base] = 1

        # Within each sheet, if multiple tables share the same title (ignoring
        # whitespace differences like "TABLE: D-3" vs "TABLE : D-3"), keep the
        # first one's title unchanged and suffix later ones with (2), (3)…
        seen: Dict[Tuple[str, str], int] = {}  # (sheet, normalised_title) -> count seen so far
        for tbl in all_tables:
            norm_key = (tbl["sheet"], re.sub(r'\s+', '', tbl.get("title", "")).lower())
            count = seen.get(norm_key, 0) + 1
            seen[norm_key] = count
            if count > 1:
                tbl["title"] = f"{tbl['title']} ({count})"

        return all_tables

    # ── Sheet processing ──────────────────────────────────────────────────────

    def _process_sheet(self, ws, sheet_name: str, filename: str) -> List[Dict]:
        grid = self._filled_grid(ws)
        if not grid:
            return []

        blocks = self._find_blocks(grid)
        print(f"[DEBUG] sheet={sheet_name!r}  grid_rows={len(grid)}  blocks={blocks}")

        results = []
        for idx, (start, end) in enumerate(blocks):
            try:
                tbl = self._extract_table(grid, start, end, sheet_name, filename, idx)
                if tbl and tbl["row_count"] > 0:
                    results.append(tbl)
            except Exception as e:
                print(f"[{sheet_name}] block {start}-{end} error: {e}")
            if idx < len(blocks) - 1:
                time.sleep(3)
        return results

    # ── Grid builder ──────────────────────────────────────────────────────────

    def _filled_grid(self, ws) -> List[List[Any]]:
        merge_lookup: Dict[Tuple[int, int], Any] = {}
        for mrng in ws.merged_cells.ranges:
            master = ws.cell(mrng.min_row, mrng.min_col).value
            for r in range(mrng.min_row, mrng.max_row + 1):
                for c in range(mrng.min_col, mrng.max_col + 1):
                    if r != mrng.min_row or c != mrng.min_col:
                        merge_lookup[(r, c)] = master

        max_row = max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                v = merge_lookup.get((cell.row, cell.column), cell.value)
                if v is not None and str(v).strip():
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        if not max_row:
            return []
        return [
            [merge_lookup.get((r, c), ws.cell(r, c).value) for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)
        ]

    # ── Block detection ───────────────────────────────────────────────────────

    @staticmethod
    def _blank(row: List[Any]) -> bool:
        return all(v is None or str(v).strip() == "" for v in row)

    @staticmethod
    def _has_table_marker(row: List[Any]) -> bool:
        text = " ".join(str(v) for v in row if v is not None)
        return bool(re.search(r"\bTABLE[\s:\-]", text, re.IGNORECASE))

    def _count_preceding_blanks(self, grid: List[List[Any]], row_idx: int) -> int:
        count = 0
        for j in range(row_idx - 1, -1, -1):
            if self._blank(grid[j]):
                count += 1
            else:
                break
        return count

    def _find_blocks(self, grid: List[List[Any]]) -> List[Tuple[int, int]]:
        # Candidate marker rows: any row containing "TABLE <letter/digit>"
        all_markers = [i for i, r in enumerate(grid) if self._has_table_marker(r)]

        if all_markers:
            first_nonblank = next((i for i, r in enumerate(grid) if not self._blank(r)), -1)
            # A marker is a genuine new-block start only when it is:
            #   (a) the first non-blank row in the sheet, OR
            #   (b) preceded by ≥2 consecutive blank rows
            markers = [
                mk for mk in all_markers
                if mk == first_nonblank or self._count_preceding_blanks(grid, mk) >= 2
            ]

            if markers:
                blocks = []
                for j, start in enumerate(markers):
                    limit = markers[j + 1] if j + 1 < len(markers) else len(grid)
                    end = limit - 1
                    while end > start and self._blank(grid[end]):
                        end -= 1
                    if end > start:
                        blocks.append((start, end))
                return blocks

        # Fallback: blank-row based detection (sheets without TABLE markers)
        blocks, current, blanks = [], None, 0
        for i, row in enumerate(grid):
            if self._blank(row):
                blanks += 1
                if blanks >= 2 and current is not None:
                    end = i - blanks
                    if end > current:
                        blocks.append((current, end))
                    current, blanks = None, 0
            else:
                blanks = 0
                if current is None:
                    current = i
        if current is not None:
            end = len(grid) - 1
            while end > current and self._blank(grid[end]):
                end -= 1
            if end > current:
                blocks.append((current, end))
        return blocks

    # ── Table extraction ──────────────────────────────────────────────────────

    def _extract_table(
        self, grid, start: int, end: int, sheet_name: str, filename: str, idx: int
    ) -> Optional[Dict]:
        block = grid[start : end + 1]
        title, description, body_start = self._strip_title_desc(block)
        body = block[body_start:]
        if not body:
            return None

        n_cols = max(len(r) for r in body)

        try:
            structure = self._direct_llm_structure(body, title, description, n_cols)
        except Exception as e:
            print(f"Structure analysis failed ({e}), using heuristic")
            structure = self._heuristic_structure(body, n_cols)

        header_rows: int = structure.get("header_rows", 1)
        skip_set: set = set(structure.get("skip_rows", []))
        columns: List[str] = list(structure.get("columns", []))

        while len(columns) < n_cols:
            columns.append(f"Col_{len(columns)+1}")
        columns = columns[:n_cols]

        # Deduplicate column names
        seen: Dict[str, int] = {}
        deduped = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                deduped.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                deduped.append(col)
        columns = deduped

        rows = []
        for i, row in enumerate(body[header_rows:], start=header_rows):
            if i in skip_set or self._blank(row):
                continue
            padded = list(row) + [None] * (n_cols - len(row))
            row_dict = {}
            for j, col in enumerate(columns):
                v = padded[j]
                if hasattr(v, "item"):
                    v = v.item()
                row_dict[col] = v
            rows.append(row_dict)

        def _ser(row):
            out = []
            for v in row:
                if v is None:
                    out.append(None)
                elif hasattr(v, "item"):
                    out.append(v.item())
                else:
                    out.append(str(v).strip())
            return out

        # Raw header rows before AI flattening (multi-level structure)
        raw_header_rows = [_ser(row) for row in body[:header_rows]]

        # Column-index rows like (1)(2)(3) that were skipped
        raw_col_num_rows = [_ser(body[i]) for i in sorted(skip_set) if i < len(body)]

        # Footer notes: text-heavy rows at the end of the body after data
        raw_notes: List[str] = []
        for row in reversed(body[header_rows:]):
            non_none = [v for v in row if v is not None and str(v).strip()]
            if not non_none:
                continue
            nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if nums > len(non_none) * 0.3:
                break  # hit a data row — stop scanning
            text = " ".join(str(v).strip() for v in non_none)
            if len(text) > 8:
                raw_notes.insert(0, text)
            if len(raw_notes) >= 10:
                break

        return {
            "id": f"{filename}__{sheet_name}__{idx}",
            "title": title or f"Table {idx+1}",
            "description": description,
            "sheet": sheet_name,
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "raw_header_rows": raw_header_rows,
            "raw_col_num_rows": raw_col_num_rows,
            "raw_notes": raw_notes,
        }

    # ── Title / description extraction ────────────────────────────────────────

    def _strip_title_desc(self, block: List[List[Any]]) -> Tuple[str, str, int]:
        def row_text(row):
            seen_vals, parts = set(), []
            for v in row:
                if v is None:
                    continue
                sv = str(v).strip()
                if sv and sv not in seen_vals:
                    seen_vals.add(sv)
                    parts.append(sv)
            return " ".join(parts)

        non_blank = [(i, r) for i, r in enumerate(block) if not self._blank(r)]
        title = description = ""
        body_start = 0
        if not non_blank:
            return title, description, body_start

        i0, r0 = non_blank[0]
        title = row_text(r0)
        body_start = i0 + 1

        if len(non_blank) > 1:
            i1, r1 = non_blank[1]
            text1 = row_text(r1)
            non_none = [v for v in r1 if v is not None]
            n_nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if text1 and len(text1) > 10 and n_nums < len(non_none) / 2:
                description = text1
                body_start = i1 + 1

        while body_start < len(block) and self._blank(block[body_start]):
            body_start += 1
        return title, description, body_start

    # ─────────────────────────────────────────────────────────────────────────
    # STRATEGY A — Direct LLM (single prompt → single response)
    # ─────────────────────────────────────────────────────────────────────────

    def _direct_llm_structure(
        self, body: List[List[Any]], title: str, description: str, n_cols: int
    ) -> Dict:
        n_sample = min(10, len(body))
        lines = []
        for i, row in enumerate(body[:n_sample]):
            vals = "\t".join("" if v is None else str(v).strip() for v in row)
            lines.append(f"Row {i+1}:\t{vals}")
        grid_text = "\n".join(lines)

        prompt = f"""Analyze this government statistical table body and return its structure.

Table: {title}
{f"Description: {description}" if description else ""}
Columns: {n_cols}, Total rows: {len(body)}

First {n_sample} rows (tab-separated):
{grid_text}

Return JSON with:
- "header_rows": count of rows at start that are column headers (not data)
- "skip_rows": 0-based row indices to skip (e.g. column-number rows like "(1)(2)(3)")
- "columns": exactly {n_cols} flat column names

For multi-level headers (group row + sub-column row), combine: "AGE_<1", "AGE_1-4"
Return ONLY valid JSON, no markdown fences."""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        text = re.sub(r"```[a-z]*\n?", "", text).strip().rstrip("`").strip()
        return json.loads(text)

    # ── LLM-based category metadata extraction ───────────────────────────────

    def extract_category_metadata(
        self,
        title: str,
        description: str,
        raw_header_rows: List[List],
        columns: List[str],
        sample_rows: List[Dict],
        raw_notes: List[str],
    ) -> List[Dict]:
        """
        Ask the LLM to semantically identify all categorical dimensions present in
        the table — e.g. Area Type (Rural/Urban), Care Setting (Institution/Domiciliary),
        Gender (Male/Female), Geography (State, District), Age Group, etc.
        Returns a list of category dicts: {name, description, values:[{value,code,description}]}
        """
        # Format original header matrix so the LLM sees the raw multi-level structure
        header_lines = []
        for i, row in enumerate(raw_header_rows):
            deduped: List[str] = []
            seen_in_row: set = set()
            for v in row:
                sv = str(v).strip() if v is not None else ""
                if sv and sv not in seen_in_row:
                    deduped.append(sv)
                    seen_in_row.add(sv)
            if deduped:
                header_lines.append(f"  Header level {i+1}: {' | '.join(deduped)}")

        # Sample the first few data rows so the LLM can see row-dimension values
        data_sample_lines = []
        for row in sample_rows[:6]:
            vals = " | ".join(str(v) for v in list(row.values())[:6] if v is not None)
            if vals:
                data_sample_lines.append(f"  {vals}")

        prompt = f"""You are analyzing a government statistical table to build a metadata catalogue.

Table: {title}
{f'Description: {description}' if description else ''}

Original multi-level column headers (before flattening):
{chr(10).join(header_lines) if header_lines else '  (single-level headers)'}

Flat column names after AI extraction ({len(columns)} total):
  {', '.join(columns[:30])}{'…' if len(columns) > 30 else ''}

First few data rows (first 6 values each):
{chr(10).join(data_sample_lines) if data_sample_lines else '  (no sample)'}
{f"Source notes: {'; '.join(raw_notes[:4])}" if raw_notes else ''}

Your task: identify EVERY categorical dimension in this table.
A "category" is any grouping variable — examples:
- Column-group headers like RURAL / URBAN (spans multiple sub-columns)
- Sub-group headers like INSTITUTION / DOMICILIARY
- Cross-tabulation variables like Male / Female, Age bands, Year
- Row-dimension labels like State, District, Block, Taluk

Rules:
- Each category name must be UNIQUE — do not repeat the same category twice
- Each value within a category must be UNIQUE — list each value only once
- Merge overlapping categories into one (e.g. "Area" and "Location Type" are the same)
- Omit "Total" or "Grand Total" rows — those are aggregates, not categories

For each category:
- Give it a clear human-readable name (e.g. "Area Type", "Care Setting", "Gender")
- Write a one-sentence description of what it classifies
- List every unique value you can identify from the headers and data sample
- Generate a short UPPERCASE_UNDERSCORE code for each value
- Write a brief plain-English description for each value

Return ONLY a valid JSON object — no prose, no markdown fences:
{{
  "categories": [
    {{
      "name": "Area Type",
      "description": "Classifies data by geographic area type",
      "values": [
        {{"value": "Rural", "code": "RURAL", "description": "Data from rural areas outside municipal limits"}},
        {{"value": "Urban", "code": "URBAN", "description": "Data from urban areas within municipal limits"}}
      ]
    }}
  ]
}}"""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        parsed = _extract_json_with_key(text, "categories")
        if parsed is None:
            return []
        return parsed.get("categories", [])

    def group_tables_by_similarity(self, tables_meta: list) -> list:
        """Use Claude to cluster tables by semantic similarity of titles/topics."""
        if not tables_meta:
            return []
        if len(tables_meta) == 1:
            return [{"name": tables_meta[0].get("title", "Table 1"), "table_ids": [tables_meta[0]["id"]]}]

        # Use simple numeric indices in the prompt — avoids Claude mangling complex ID strings.
        # We map indices back to real IDs after parsing.
        index_to_id = {str(i + 1): t["id"] for i, t in enumerate(tables_meta)}

        lines = "\n".join(
            f"{i+1}. title={t.get('title', '')!r}  sheet={t.get('sheet', '')!r}  desc={t.get('description', '')!r}"
            for i, t in enumerate(tables_meta)
        )
        prompt = f"""You are given tables extracted from an Excel file. Each table has a number (1, 2, 3 …).

Tables:
{lines}

Group tables that cover the same topic but different breakdowns (e.g. rural/urban, male/female, district-wise, age-group).
Example: "Still Birth Rural" and "Still Birth Urban" → one group called "Still Birth".

Return ONLY valid JSON, no prose, no markdown:
{{"groups": [{{"name": "Group Name", "indices": [1, 2]}}, ...]}}

Rules:
- Use the table numbers (integers) in "indices", not titles.
- Every table number from 1 to {len(tables_meta)} must appear in exactly one group.
- Use a short descriptive group name (e.g. "Still Birth", "Live Birth", "Death Registration").
- Tables on clearly different topics go in separate groups."""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        parsed = _extract_json_with_key(text, "groups")
        if parsed:
            groups = parsed.get("groups", [])
            result = []
            seen_indices = set()
            for g in groups:
                ids = []
                for idx in g.get("indices", []):
                    key = str(idx)
                    if key in index_to_id and key not in seen_indices:
                        ids.append(index_to_id[key])
                        seen_indices.add(key)
                if ids:
                    result.append({"name": g.get("name", "Group"), "table_ids": ids})
            # append any tables Claude forgot to include
            for key, tid in index_to_id.items():
                if key not in seen_indices:
                    result.append({"name": f"Table {key}", "table_ids": [tid]})
            if result:
                return result

        return [{"name": t.get("title", f"Table {i+1}"), "table_ids": [t["id"]]}
                for i, t in enumerate(tables_meta)]

    # ── DES Catalogue enrichment ──────────────────────────────────────────────

    def enrich_for_catalogue(self, table: dict) -> dict:
        """Use Claude to generate DES-catalogue-compatible metadata for one table."""
        columns = table.get("columns", [])
        sample_rows = table.get("rows", [])[:6]

        prompt = f"""You are a data cataloguer for the Delhi Economic Survey (DES). Given an extracted Excel table, generate structured catalogue metadata.

Table Title: {table.get('title', '')}
Sheet: {table.get('sheet', '')}
Columns ({len(columns)}): {columns}
Sample rows (first 6): {json.dumps(sample_rows, default=str)}
Existing description: {table.get('description', '')}

Return ONLY a JSON object with exactly these fields:
{{
  "short_description": "1-2 sentence factual summary of what data this table contains",
  "long_description": "3-5 sentence detailed description: what is measured, geographic scope, time period, disaggregation levels, and how to interpret the data",
  "units": "measurement unit as a short string, e.g. 'Count', 'Percentage', 'Rs. Crore', 'Lakhs', 'Rate per 1000'",
  "classifications": {{
    "dimension_name": ["value1", "value2"]
  }},
  "age_column_keys": {{
    "col_name_or_key": "readable age label"
  }}
}}

Rules for classifications:
- Identify categorical dimensions from column names and sample data values (e.g. gender_classification, area_type, administrative_body, place_of_occurrence)
- For each dimension list up to 15 unique values found in column names or sample data
- Use snake_case dimension names
- Do NOT include numeric/quantitative columns as classifications
- If no clear categorical dimensions exist, return an empty object {{}}

Rules for age_column_keys:
- Only populate if the table has columns representing age groups (e.g. "<1", "1-4", "15-24", "65+")
- Map the column name to a readable human label
- Return empty object {{}} if no age-related columns exist

Return only valid JSON, no markdown fences, no explanation."""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        # strip markdown fences
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text).strip()
        try:
            result = json.loads(text)
        except json.JSONDecodeError:
            parsed = _extract_json_with_key(text, "short_description")
            result = parsed or {}

        return {
            "short_description": result.get("short_description", table.get("description", "")),
            "long_description": result.get("long_description", table.get("description", "")),
            "units": result.get("units", "Count"),
            "classifications": result.get("classifications", {}),
            "age_column_keys": result.get("age_column_keys", {}),
        }

    # ── Heuristic fallback ────────────────────────────────────────────────────

    @staticmethod
    def _heuristic_structure(body: List[List[Any]], n_cols: int) -> Dict:
        if not body:
            return {"header_rows": 0, "skip_rows": [], "columns": [f"Col_{i+1}" for i in range(n_cols)]}
        skip_rows = []
        for i, row in enumerate(body[:6]):
            non_none = [v for v in row if v is not None]
            if not non_none:
                continue
            col_nums = sum(1 for v in non_none if re.match(r"^\(\d+\)$", str(v).strip()))
            if col_nums > len(non_none) * 0.5:
                skip_rows.append(i)
        header_rows = 1
        for i, row in enumerate(body[:5]):
            non_none = [v for v in row if v is not None]
            nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if nums > len(non_none) * 0.4 and i > 0:
                header_rows = i
                break
        columns = [str(v).strip() if v is not None else f"Col_{i+1}" for i, v in enumerate(body[0])]
        while len(columns) < n_cols:
            columns.append(f"Col_{len(columns)+1}")
        return {"header_rows": header_rows, "skip_rows": skip_rows, "columns": columns}
