"""Parses DES metadata workbooks (catalogue_summary / dataset_inventory_list /
classification sheets / concept sheet) into structured data used to
(a) prefill the Create Metadata form and (b) automatically match uploaded
metadata files against extracted dataset tables.
"""

from io import BytesIO

import openpyxl

FREQUENCY_MAP = {
    "annual": "Annual",
    "annually": "Annual",
    "monthly": "Monthly",
    "quarterly": "Quarterly",
    "decennial": "Decennial",
    "ad-hoc": "Ad-hoc",
    "ad hoc": "Ad-hoc",
    "adhoc": "Ad-hoc",
}

FIXED_SHEETS = {"catalogue_summary", "dataset_inventory_list"}


def _normalize_frequency(raw: str) -> str:
    if not raw:
        return ""
    return FREQUENCY_MAP.get(str(raw).strip().lower(), "")


def _rows(ws):
    return [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]


def _rows_to_dicts(rows):
    if len(rows) < 2:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {header[i]: row[i] for i in range(len(header)) if header[i]}
        out.append(d)
    return out


def _find_header_row_index(rows, must_equal: str):
    needle = must_equal.strip().lower()
    for i, row in enumerate(rows[:10]):
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() == needle:
                return i
    return -1


def _get(record: dict, key: str) -> str:
    val = record.get(key)
    return str(val).strip() if val is not None else ""


def _parse_catalogue_summary_from_wb(wb) -> dict:
    if "catalogue_summary" not in wb.sheetnames:
        raise ValueError("No 'catalogue_summary' sheet found in this workbook")

    rows = _rows(wb["catalogue_summary"])
    if len(rows) < 2:
        raise ValueError("'catalogue_summary' sheet has no data row")

    record = _rows_to_dicts(rows)[0]
    product = _get(record, "product")

    return {
        "title": product,
        "product": product,
        "category": _get(record, "category"),
        "geography": _get(record, "geography"),
        "frequency": _normalize_frequency(_get(record, "frequency")),
        "time_period": _get(record, "time_period"),
        "data_source": _get(record, "data_source"),
        "description": _get(record, "description"),
        "last_updated": _get(record, "last_updated_date"),
        "future_release": _get(record, "future_release"),
        "key_statistics": _get(record, "key_statistics"),
        "remarks": _get(record, "remarks"),
    }


def parse_catalogue_summary(file_bytes: bytes) -> dict:
    """Group-level fields used to prefill the Create Metadata form."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    return _parse_catalogue_summary_from_wb(wb)


def parse_inventory(wb) -> list:
    if "dataset_inventory_list" not in wb.sheetnames:
        return []
    rows = _rows(wb["dataset_inventory_list"])
    out = []
    for r in _rows_to_dicts(rows):
        unique_id = r.get("unique dataset id")
        if not unique_id:
            continue
        out.append({
            "sl_no": r.get("sl#"),
            "unique_dataset_id": str(unique_id).strip(),
            "table_id": str(r.get("table id") or "").strip(),
            "short_description": str(r.get("dataset short description") or "").strip(),
            "long_description": str(r.get("dataset long description") or "").strip(),
        })
    return out


def parse_concepts(wb) -> list:
    sheet_name = next((n for n in wb.sheetnames if n.lower() == "nmds_concept_meta_data"), None)
    if not sheet_name:
        return []
    rows = _rows(wb[sheet_name])
    out = []
    for r in _rows_to_dicts(rows):
        concept = r.get("Concept Name")
        if not concept:
            continue
        out.append({
            "item_no": r.get("Item No"),
            "concept": concept,
            "details": r.get("Details (Summary)"),
        })
    return out


def parse_classifications(wb) -> dict:
    skip = FIXED_SHEETS | {"nmds_concept_meta_data"}
    classifications = {}
    for name in wb.sheetnames:
        if name.lower() in skip:
            continue
        rows = _rows(wb[name])
        header_idx = _find_header_row_index(rows, "code")
        if header_idx == -1:
            continue
        header = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
        entries = []
        for row in rows[header_idx + 1:]:
            d = {header[i]: row[i] for i in range(len(header)) if header[i]}
            if d.get("code") in (None, ""):
                continue
            entries.append({"code": d.get("code"), "value": d.get("value"), "definition": d.get("definition")})
        if entries:
            classifications[name] = entries
    return classifications


def parse_metadata_workbook(file_bytes: bytes, filename: str = "") -> dict:
    """Full parse of a metadata workbook: summary + inventory + concepts +
    classifications. Used by the batch auto-mapping flow."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    summary = _parse_catalogue_summary_from_wb(wb)
    return {
        "file_name": filename or "",
        "summary": summary,
        "inventory": parse_inventory(wb),
        "concepts": parse_concepts(wb),
        "classifications": parse_classifications(wb),
    }
